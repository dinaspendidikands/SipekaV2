"""Generator data dummy SIPEKA -> DUMMY-DATA-SIPEKA.xlsx
Kolom & agregat dibuat persis seperti setup()/submit di apps-script/Code.gs."""
import hashlib, random, uuid, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(2026)
PASSWORD = "sipeka2026"
HASH = hashlib.sha256(PASSWORD.encode("utf-8")).hexdigest()
TXT = "@"  # number_format teks (jaga NIP/NPSN/HP/0 di depan)

# ---------- struktur indikator (salin dari Code.gs) ----------
KS_SDSMP = [
    ("Performance Sekolah", [("1.1","Kebersihan Lingkungan Satuan Pendidikan"),("1.2","Kebersihan Lingkungan Belajar"),("1.3","Kelengkapan Sanitasi dan Kebersihan Kamar Mandi"),("1.4","Kerindangan dan Apotik Hidup Sekolah"),("1.5","Mewujudkan Sekolah Adiwiyata"),("1.6","Mewujudkan sekolah dan kantin Sehat")]),
    ("Tata Kelola Sampah", [("2.1","Ketersediaan Tong Sampah Sesuai Peruntukan"),("2.2","Pengelolaan bank Sampah"),("2.3","Tempat Pengomposan dan Pengolahan Sampah Organik"),("2.4","Produk Daur Ulang")]),
    ("Inovasi Sekolah", [("3.1","Jumlah Inovasi di Sekolah"),("3.2","Pemanfaatan Teknologi dalam Pembelajaran"),("3.3","Upaya meningkatkan literasi dan numerasi")]),
    ("Indeks Kepuasan & Kepemimpinan", [("4.1","Indeks kepuasan Orang Tua Murid"),("4.2","Indeks kepuasan Siswa"),("4.3","Indeks Kepuasan guru dan staff")]),
    ("Prestasi Murid", [("5.1","Jumlah murid dalam mengikuti perlombaan di bidang akademik dan non akademik"),("5.2","Jumlah Murid yang Berprestasi di bidang akademik"),("5.3","Jumlah Murid yang Berprestasi di bidang non akademik"),("5.4","Jumlah Murid yang tidak Lancar Calistung"),("5.5","Partisipasi Murid dalam Ekstrakurikuler")]),
    ("Kelengkapan Dokumen Sekolah", [("6.1","Dokumen Rencana Kerja Sekolah"),("6.2","Dokumen Kurikulum Satuan Pendidikan"),("6.3","Dokumen Inventaris Barang"),("6.4","Dokumen Administrasi Murid"),("6.5","Laporan Penggunaan Dana BOS")]),
]
KS_PAUD = [
    ("Performance", [("1.1","Kebersihan Lingkungan Satuan Pendidikan"),("1.2","Kebersihan Lingkungan Belajar"),("1.3","Kelengkapan Sanitasi dan Kebersihan Kamar Mandi"),("1.4","Kerindangan dan Apotik Hidup Sekolah")]),
    ("Kebijakan Sekolah Ramah Anak (SRA)", [("2.1","Area bermain aman bagi anak"),("2.2","Terpisah toilet laki-laki dan perempuan"),("2.3","Parenting")]),
    ("Pemanfaatan Teknologi & 7 KAIH", [("3.1","Pemanfaatan Teknologi dan tersedianya APE dalam pembelajaran"),("3.2","Pelaksanaan kegiatan 7 KAIH")]),
    ("Data Tumbuh Kembang Anak", [("4.1","Pertumbuhan dan perkembangan anak")]),
    ("Kemitraan", [("5.1","Kerjasama dengan puskesmas, posyandu, atau lembaga lain"),("5.2","Keterlibatan orang tua dalam kegiatan sekolah")]),
    ("Dokumen Sekolah", [("6.1","Dokumen Rencana Kerja Sekolah"),("6.2","Dokumen Kurikulum Satuan Pendidikan"),("6.3","Dokumen Inventaris Barang"),("6.4","Dokumen Administrasi Murid"),("6.5","Laporan Penggunaan Dana BOS")]),
]
GURU_GRP = [("Manajemen Ruang & Lingkungan Sekolah",7),("Tata Kelola Sampah",4),("Inovasi & Pengembangan Profesi",8),("Kepuasan, Komunikasi & Prestasi Siswa",10),("Praktek Pembelajaran-Perangkat Pembelajaran",4)]
GURU_SKOR = {3: 7}  # kelompok index 3 (Kepuasan...): 7 butir pertama = survey desimal 0-1

BULAN = "Mei"; MINGGU = "II"; TW = "TW2"
def ts(d): return datetime.datetime(2026,5,d,9,0,0) + datetime.timedelta(minutes=random.randint(0,400))

# ---------- master entitas (konsisten antar sheet) ----------
KEC = ["Lubuk Pakam","Tanjung Morawa","Pancur Batu"]
PENGAWAS = [
    ("196508121990031005","Drs. Suparman, M.Pd","081361110001"),
    ("197102201995121002","Budi Santoso, M.Pd","081361110002"),
    ("197505102000031003","Hendra Gunawan, S.Pd, M.Pd","081361110003"),
    ("196803151990032004","Hj. Rosmawati, S.Pd","081361110004"),
]
ADMIN_KEC = [("Lubuk Pakam","Andi Pratama","081362220001"),("Tanjung Morawa","Citra Dewi","081362220002"),("Pancur Batu","Eko Nugroho","081362220003")]
# sekolah: (npsn, nama, jenjang, kec, ks, hpKS, pengawas_nama, hpPengawas)
P = {n[1]: n[2] for n in PENGAWAS}
SEK = [
    ("10212001","TK Negeri Pembina Lubuk Pakam","TK","Lubuk Pakam","Sri Wahyuni, S.Pd","081370010001","Hj. Rosmawati, S.Pd"),
    ("10212002","PAUD Tunas Bangsa","PAUD","Tanjung Morawa","Nurhayati, S.Pd","081370010002","Hj. Rosmawati, S.Pd"),
    ("10212003","TK Aisyiyah Bustanul Athfal","TK","Pancur Batu","Dewi Sartika, S.Pd","081370010003","Hj. Rosmawati, S.Pd"),
    ("10212004","TK Kasih Ibu","TK","Lubuk Pakam","Lestari Ningsih, S.Pd","081370010004","Hj. Rosmawati, S.Pd"),
    ("10212101","SD Negeri 101878 Lubuk Pakam","SD","Lubuk Pakam","Drs. Ahmad Yani","081370011001","Drs. Suparman, M.Pd"),
    ("10212102","SD Negeri 105292 Tanjung Morawa","SD","Tanjung Morawa","Siti Aminah, S.Pd","081370011002","Drs. Suparman, M.Pd"),
    ("10212103","SD Swasta Methodist","SD","Pancur Batu","Marolop Sihombing, S.Pd","081370011003","Budi Santoso, M.Pd"),
    ("10212104","SD Negeri 107400 Pancur Batu","SD","Pancur Batu","Rohani Br Tarigan, S.Pd","081370011004","Budi Santoso, M.Pd"),
    ("10212201","SMP Negeri 1 Lubuk Pakam","SMP","Lubuk Pakam","Drs. Parlindungan","081370012001","Hendra Gunawan, S.Pd, M.Pd"),
    ("10212202","SMP Negeri 2 Tanjung Morawa","SMP","Tanjung Morawa","Erni Susanti, S.Pd","081370012002","Hendra Gunawan, S.Pd, M.Pd"),
    ("10212203","SMP Swasta Budi Mulia","SMP","Pancur Batu","Jhonson Simbolon, S.Pd","081370012003","Hendra Gunawan, S.Pd, M.Pd"),
    ("10212204","SMP Negeri 3 Lubuk Pakam","SMP","Lubuk Pakam","Maria Magdalena, S.Pd","081370012004","Hendra Gunawan, S.Pd, M.Pd"),
]
def hp_pengawas(nama): return next(p[2] for p in PENGAWAS if p[1]==nama)

# ---------- styling ----------
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor="2E5A87")
BODY_FONT = Font(name="Arial", size=10)

def write_sheet(wb, name, headers, rows, text_cols):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c); cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    # format teks utk kolom NIP/NPSN/HP agar 0 di depan & angka panjang tak rusak
    for ci in text_cols:
        for ri in range(2, len(rows)+2):
            ws.cell(ri, ci+1).number_format = TXT
    ws.freeze_panes = "A2"
    for c in range(1, len(headers)+1):
        ws.column_dimensions[ws.cell(1,c).column_letter].width = 16
    return ws

wb = Workbook(); wb.remove(wb.active)

# ===== AKUN_PENGAWAS =====
hdr = ["NIP (USERNAME)","NAMA PENGAWAS","HP","HASH PASSWORD","KUNCI","WAJIB GANTI","TERAKHIR LOGIN"]
rows = [[nip, nama, hp, HASH, str(uuid.uuid4()), "T", ""] for nip,nama,hp in PENGAWAS]
write_sheet(wb, "AKUN_PENGAWAS", hdr, rows, text_cols=[0,2])

# ===== AKUN_KECAMATAN =====
hdr = ["KECAMATAN (USERNAME)","NAMA PETUGAS","HP","HASH PASSWORD","KUNCI","WAJIB GANTI","TERAKHIR LOGIN"]
rows = [[kec, nama, hp, HASH, str(uuid.uuid4()), "T", ""] for kec,nama,hp in ADMIN_KEC]
write_sheet(wb, "AKUN_KECAMATAN", hdr, rows, text_cols=[2])

# ===== header builder KS =====
def header_ks(grp, paud):
    h = ["Timestamp","Email Address","NPSN","Kecamatan","JENJANG SEKOLAH","NAMA SEKOLAH","NAMA PENGAWAS","Nomor HP Pengawas","NAMA KEPALA SEKOLAH","Nomor HP Kepala Sekolah"]
    if not paud: h.append("Periode Penilaian")
    h += ["Bulan","Minggu"]
    for _,inds in grp:
        for kode,nama in inds:
            h += [f"{kode} {nama} [Baseline]", f"{kode} {nama} [Target]", f"{kode} {nama} [Capaian]", f"Link Dokumentasi {kode} {nama}"]
    h += ["Catatan dan Rekomendasi"]
    if not paud: h += ["Nomor Urut"]
    h += ["Rata Baseline","Rata Target","Rata Capaian","Kinerja Keseluruhan"]
    if paud:
        for nama,_ in grp: h += [f"Rata Baseline {nama}",f"Rata Target {nama}",f"Rata Capaian {nama}",f"Kinerja {nama}"]
    else:
        for i,(nama,_) in enumerate(grp): h += [f"BASELINE {i+1}",f"TARGET {i+1}",f"CAPAIAN {i+1}",f"Kinerja {nama}"]
    h += ["ID KIRIM"]
    return h

def row_ks(s, grp, paud):
    npsn,nama_sek,jenjang,kec,ks,hpKS,peng = s
    hpP = hp_pengawas(peng)
    row = [ts(random.randint(20,28)), "pengawas@dinas.belajar.id", npsn, kec, jenjang, nama_sek, peng, hpP, ks, hpKS]
    if not paud: row.append(TW)
    row += [BULAN, MINGGU]
    jml = sum(len(i) for _,i in grp)
    sB=sT=sC=0; gstat=[]
    for _,inds in grp:
        gb=gt=gc=0
        for _ in inds:
            b=random.randint(2,4); c=random.randint(b,5); t=max(c,random.randint(b,5))
            row += [b,t,c,"https://drive.google.com/file/d/contoh"]
            sB+=b; sT+=t; sC+=c; gb+=b; gt+=t; gc+=c
        gstat.append((gb,gt,gc,len(inds)))
    row.append("Pertahankan kebersihan dan tingkatkan inovasi pembelajaran.")
    if not paud: row.append("")
    row += [sB/jml, sT/jml, sC/jml, (sC-sB)/jml]
    for gb,gt,gc,n in gstat:
        row += [gb/n, gt/n, gc/n, (gc-gb)/n]
    row.append(str(uuid.uuid4()))
    return row

# ===== KS_PAUD =====
paud_sek = [s for s in SEK if s[2] in ("TK","PAUD")]
write_sheet(wb, "KS_PAUD", header_ks(KS_PAUD, True), [row_ks(s, KS_PAUD, True) for s in paud_sek], text_cols=[2,7,9])

# ===== KS_SD_SMP =====
sdsmp_sek = [s for s in SEK if s[2] in ("SD","SMP")]
write_sheet(wb, "KS_SD_SMP", header_ks(KS_SDSMP, False), [row_ks(s, KS_SDSMP, False) for s in sdsmp_sek], text_cols=[2,7,9])

# ===== GURU =====
def header_guru():
    h = ["Timestamp","Email Address","KECAMATAN","JENJANG","NPSN","SATUAN PENDIDIKAN","NAMA KEPALA SEKOLAH","NOMOR HP KEPALA SEKOLAH","NAMA GURU","STATUS","PANGKAT/GOL.RUANG","NIP/NUPTK","MATA PELAJARAN","NOMOR HP GURU","BULAN"]
    for gi,(nama,cnt) in enumerate(GURU_GRP):
        for b in range(1,cnt+1): h.append(f"Penilaian {nama} [Butir {b}]")
        h.append(f"Link Bukti Dukung Indikator {gi+1}. {nama}")
    for nama,_ in GURU_GRP: h.append(f"Rata Capaian {nama}")
    h += ["Rata Keseluruhan","ID KIRIM"]
    return h

NAMA_GURU = ["Agus Salim","Budiman","Cut Ny Dien","Dewi Lestari","Eka Putri","Fauzan Akbar","Gunawan","Hesti Wulandari","Indah Permata","Joko Susilo","Kartika Sari","Lukman Hakim","Mega Wati","Nanda Pratama","Oki Setiawan","Putri Ayu","Rahmat Hidayat","Sari Bulan","Tono Wijaya","Umar Bakri","Vina Maulida","Wahyu Ramadhan","Yanti Kusuma","Zulkifli"]
STATUS = ["PNS","P3K","P3K PW","Honorer","GTY"]
PANGKAT_PNS = ["III/a","III/b","III/c","III/d","IV/a"]
MAPEL_SMP = ["Matematika","IPA","Bahasa Indonesia","Bahasa Inggris","IPS","PJOK","Seni Budaya","PPKn"]

def row_guru(s, idx):
    npsn,nama_sek,jenjang,kec,ks,hpKS,peng = s
    nama = NAMA_GURU[idx % len(NAMA_GURU)] + (f" {idx//len(NAMA_GURU)+1}" if idx>=len(NAMA_GURU) else "")
    status = random.choice(STATUS)
    pangkat = random.choice(PANGKAT_PNS) if status in ("PNS","P3K") else "-"
    mapel = "Guru Kelas" if jenjang in ("SD","TK","PAUD") else random.choice(MAPEL_SMP)
    nip = "".join(str(random.randint(0,9)) for _ in range(18))
    hp = "0813" + "".join(str(random.randint(0,9)) for _ in range(8))
    row = [ts(random.randint(20,28)), "kepsek@sekolah.belajar.id", kec, jenjang, npsn, nama_sek, ks, hpKS, nama, status, pangkat, nip, mapel, hp, BULAN]
    ratas=[]
    for gi,(gnama,cnt) in enumerate(GURU_GRP):
        nskor = GURU_SKOR.get(gi,0); s_=0
        for bi in range(cnt):
            if bi < nskor:
                v = round(random.uniform(0.70,1.0),2)
            else:
                v = 1 if random.random()<0.8 else 0
            row.append(v); s_+=v
        row.append("https://drive.google.com/file/d/contoh")
        ratas.append(round(s_/cnt*10000)/100)
    for r in ratas: row.append(r)
    row.append(round(sum(ratas)/len(ratas)*100)/100)
    row.append(str(uuid.uuid4()))
    return row

guru_rows=[]; gidx=0
for s in SEK:
    n = 3 if s[2] in ("TK","PAUD") else 4
    for _ in range(n):
        guru_rows.append(row_guru(s, gidx)); gidx+=1
write_sheet(wb, "GURU", header_guru(), guru_rows, text_cols=[4,7,11,13])

wb.save("DUMMY-DATA-SIPEKA.xlsx")
print("OK", {ws.title: ws.max_row-1 for ws in wb.worksheets})
print("Password semua akun:", PASSWORD)
print("HASH:", HASH)
